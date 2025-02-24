import os
import websocket
import openpyxl
import datetime
import hashlib
import base64
import hmac
import json
from urllib.parse import urlencode
import time
import ssl
from wsgiref.handlers import format_date_time
from datetime import datetime
from time import mktime
import _thread as thread
import csv
import xml.etree.ElementTree as ET
# -*- coding: utf-8 -*-

STATUS_FIRST_FRAME = 0  # 第一帧的标识
STATUS_CONTINUE_FRAME = 1  # 中间帧标识
STATUS_LAST_FRAME = 2  # 最后一帧的标识

#  BusinessArgs参数常量
SUB = "ise"
ENT = "cn_vip"
CATEGORY = "read_sentence"

# 读取音频-文本对照表
def read_audio_text_mapping():
    audio_text_mapping = {}
    # workbook = openpyxl.load_workbook('src/main/java/fun/redamancyxun/chinese/backend/score/AudioTextMapping.xlsx')
    workbook = openpyxl.load_workbook('/score/AudioTextMapping.xlsx')
    sheet = workbook.active

    # 跳过表头
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pcm_path, text = row
        audio_text_mapping[pcm_path] = text

    return audio_text_mapping

class Ws_Param(object):
    # 初始化
    def __init__(self, APPID, APIKey, APISecret, AudioFile, Text):
        self.APPID = APPID
        self.APIKey = APIKey
        self.APISecret = APISecret
        self.AudioFile = AudioFile
        self.Text = Text

        # 公共参数(common)
        self.CommonArgs = {"app_id": self.APPID}
        # 业务参数(business)，更多个性化参数可在官网查看
        self.BusinessArgs = {"category": CATEGORY, "sub": SUB, "ent": ENT, "cmd": "ssb", "auf": "audio/L16;rate=16000",
                             "aue": "raw", "text": self.Text, "ttp_skip": True, "aus": 1}

    # 生成url
    def create_url(self):
        url = 'ws://ise-api.xfyun.cn/v2/open-ise'
        now = datetime.now()
        date = format_date_time(mktime(now.timetuple()))

        signature_origin = "host: " + "ise-api.xfyun.cn" + "\n"
        signature_origin += "date: " + date + "\n"
        signature_origin += "GET " + "/v2/open-ise " + "HTTP/1.1"
        signature_sha = hmac.new(self.APISecret.encode('utf-8'), signature_origin.encode('utf-8'),
                                 digestmod=hashlib.sha256).digest()
        signature_sha = base64.b64encode(signature_sha).decode(encoding='utf-8')

        authorization_origin = "api_key=\"%s\", algorithm=\"%s\", headers=\"%s\", signature=\"%s\"" % (
            self.APIKey, "hmac-sha256", "host date request-line", signature_sha)
        authorization = base64.b64encode(authorization_origin.encode('utf-8')).decode(encoding='utf-8')

        v = {
            "authorization": authorization,
            "date": date,
            "host": "ise-api.xfyun.cn"
        }
        url = url + '?' + urlencode(v)

        return url

# 收到websocket消息的处理
def on_message(ws, message):
    try:
        code = json.loads(message)["code"]
        sid = json.loads(message)["sid"]
        if code != 0:
            errMsg = json.loads(message)["message"]
            print("sid:%s call error:%s code is:%s" % (sid, errMsg, code))
        else:
            data = json.loads(message)["data"]
            status = data["status"]
            result = data["data"]
            if (status == 2):
                xml_data = base64.b64decode(result).decode("gbk")
                root = ET.fromstring(xml_data)
                # 将XML保存到文件中，使用UTF-8编码
                pcm_filename = os.path.basename(wsParam.AudioFile)
                xml_filename = os.path.splitext(pcm_filename)[0] + '.xml'
                # xml_filepath = os.path.join('src/main/java/fun/redamancyxun/chinese/backend/score/xml', xml_filename)
                xml_filepath = os.path.join('/score/xml', xml_filename)
                with open(xml_filepath, "w", encoding="utf-8") as f:
                    f.write(xml_data)
                print(f"XML result saved as {xml_filepath}")

    except Exception as e:
        print("receive msg, but parse exception:", e)

# 收到websocket错误的处理
def on_error(ws, error):
    print("### error:", error)

# 收到websocket关闭的处理
def on_close(ws):
    print("### closed ###")

# 收到websocket连接建立的处理
def on_open(ws):
    def run(*args):
        frameSize = 1280
        intervel = 0.04
        segment_duration = 120  # 120 seconds
        bytes_per_second = 32000  # 16000 samples/sec * 2 bytes/sample
        segment_size = bytes_per_second * segment_duration  # bytes per segment

        current_char_index = 0  # 当前已匹配的字符索引

        with open(wsParam.AudioFile, "rb") as fp:
            while True:
                segment = fp.read(segment_size)
                if not segment:
                    break

                # 确保我们处理的文本长度与当前已匹配的字符相对应
                # 计算当前分段对应的文本部分
                remaining_text = wsParam.Text[current_char_index:]
                segment_text_length = len(segment) // 2  # 估算每段对应的字符数，假设每个字符占2字节
                segment_text = remaining_text[:segment_text_length]

                if not segment_text:
                    break

                # 更新业务参数中的文本
                wsParam.BusinessArgs["text"] = segment_text

                status = STATUS_FIRST_FRAME
                start_pos = 0
                while start_pos < len(segment):
                    end_pos = min(start_pos + frameSize, len(segment))
                    buf = segment[start_pos:end_pos]
                    start_pos = end_pos

                    if status == STATUS_FIRST_FRAME:
                        d = {"common": wsParam.CommonArgs,
                             "business": wsParam.BusinessArgs,
                             "data": {"status": 0}}
                        d = json.dumps(d)
                        ws.send(d)
                        status = STATUS_CONTINUE_FRAME
                    elif status == STATUS_CONTINUE_FRAME:
                        d = {"business": {"cmd": "auw", "aus": 2, "aue": "raw"},
                             "data": {"status": 1, "data": str(base64.b64encode(buf).decode())}}
                        ws.send(json.dumps(d))
                    time.sleep(intervel)

                # 发送最后一帧
                d = {"business": {"cmd": "auw", "aus": 4, "aue": "raw"},
                     "data": {"status": 2, "data": str(base64.b64encode(segment[start_pos:]).decode())}}
                ws.send(json.dumps(d))
                time.sleep(1)

                # 更新已匹配的字符索引
                current_char_index += len(segment_text)

        ws.close()

    thread.start_new_thread(run, ())


if __name__ == "__main__":
    audio_text_mapping = read_audio_text_mapping()

    for audio_file, text in audio_text_mapping.items():
        # # 尝试读取文件
        # audio_file_open = audio_file + '_16000'  # 添加后缀
        # audio_file_path = os.path.join('cn', audio_file_open + '.pcm')  # 构建完整的文件路径
        # 处理文件的逻辑
        wsParam = Ws_Param(APPID='4884d8d1', APISecret='MGJlOThkM2Q1Njc1YzcxODliMTQ4ZWZi',
                            APIKey='269cba2b9e41326b0303e1a1b36b208b',
                            AudioFile=os.path.join('pcm', audio_file), Text='\uFEFF' + text)
        if os.path.exists(os.path.join('pcm', audio_file)):
            websocket.enableTrace(False)
            wsUrl = wsParam.create_url()
            ws = websocket.WebSocketApp(wsUrl, on_message=on_message, on_error=on_error, on_close=on_close)
            ws.on_open = on_open
            ws.run_forever(sslopt={"cert_reqs": ssl.CERT_NONE})
        else:
            print(f"文件 '{os.path.join('pcm', audio_file)}' 不存在")
