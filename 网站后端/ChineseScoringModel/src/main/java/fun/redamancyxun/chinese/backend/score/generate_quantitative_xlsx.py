import os
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl

audio_text_mapping = {}
# workbook = openpyxl.load_workbook('src/main/java/fun/redamancyxun/chinese/backend/score/AudioTextMapping.xlsx')
workbook = openpyxl.load_workbook('/score/AudioTextMapping.xlsx')
sheet = workbook.active

# 跳过表头
for row in sheet.iter_rows(min_row=2, values_only=True):
    pcm_path, text = row
    audio_text_mapping[pcm_path] = text

for filename in audio_text_mapping.keys():

    filename = (os.path.splitext(filename)[0] + '.xml').replace('pcm', 'xml')

    print("正在处理" + filename)

    # 创建空的DataFrame来存储当前文件的数据
    all_data = []

    try:
        # 解析 XML 文件
        tree = ET.parse(filename)
        root = tree.getroot()

        # 遍历 XML 树
        for read_sentence in root.findall('.//read_sentence'):
            if not read_sentence.get('content'):
                continue
            read_sentence_data = {
                'filename': filename,
                'content': read_sentence.get('content'),
                # 'type': 'total',#
                'beg_pos': read_sentence.get('beg_pos'),
                'end_pos': read_sentence.get('end_pos'),
                # 'time_len': read_sentence.get('time_len'),
                'fluency_score': read_sentence.get('fluency_score'),
                'integrity_score': read_sentence.get('integrity_score'),
                'phone_score': read_sentence.get('phone_score'),
                'tone_score': read_sentence.get('tone_score'),
                'total_score': read_sentence.get('total_score')
            }
            all_data.append(read_sentence_data)

            for sentence in read_sentence.findall('.//sentence'):
                # sentence_data = {
                #     'filename': filename,
                #     'content': sentence.get('content'),
                #     'type': 'sentence',
                #     'beg_pos': sentence.get('beg_pos'),
                #     'end_pos': sentence.get('end_pos'),
                #     # 'time_len': sentence.get('time_len'),
                #     'phone_score': '',
                #     'tone_score': '',
                #     'total_score': ''
                # }
                # all_data.append(sentence_data)

                for word in sentence.findall('.//word'):
                    sylls = word.findall('.//syll')
                    for syll in sylls:
                        # syll_data = {
                        #     'filename': filename,
                        #     'content': syll.get('content'),
                        #     'type': 'syll',
                        #     'beg_pos': syll.get('beg_pos'),
                        #     'end_pos': syll.get('end_pos'),
                        #     'time_len': syll.get('time_len'),
                        #     'dp_message': syll.get('dp_message'),
                        #     'rec_node_type': syll.get('rec_node_type'),
                        #     'symbol': syll.get('symbol'),
                        # }
                        # all_data.append(syll_data)

                        phones = syll.findall('.//phone')
                        for phone in phones:
                            phone_data = {
                                'filename': filename,
                                'content': phone.get('content'),
                                # 'type': 'phone',
                                'beg_pos': phone.get('beg_pos'),
                                'end_pos': phone.get('end_pos'),
                                # 'time_len': phone.get('time_len'),
                                'dp_message': phone.get('dp_message'),
                                # 'rec_node_type': phone.get('rec_node_type'),#
                                # 'symbol': '',
                                'mono_tone': phone.get('mono_tone'),
                                'is_yun': phone.get('is_yun'),
                                'perr_msg': phone.get('perr_msg'),
                                # 'perr_level_msg': phone.get('perr_level_msg')
                            }
                            all_data.append(phone_data)

        # 将当前文件的数据存储到DataFrame中
        df = pd.DataFrame(all_data)

        # 将DataFrame保存为Excel文件，文件名与XML文件一致
        df.to_excel(filename.replace('.xml', ".xlsx").replace('xml', 'xlsx'), index=False)

    except Exception as e:
        print(f"处理文件 {filename} 时出错: {e}")